#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
STATCOLL FAM - Traitement des statistiques collaborateurs/famille (Sage 100)

- Lit un fichier Excel Sage 100 "Statistiques Collaborateurs/Famille"
- Gère la structure : Commercial sur une ligne, puis codes famille (F001, F002...)
- Produit un Excel avec 2 feuilles :
  1. "Statistiques" : Détail complet (Commercial, Code Famille, Intitulé, CA, Marge, %)
  2. "KPI" : Tableau de bord avec :
     - Résumés par Commercial
     - Résumés par Famille
     - 4 tableaux à double entrée (Commerciaux × Familles) :
       * Nbre de ventes
       * CA total
       * Marge totale
       * Profitabilité (Marge / CA)

Dépendances :
    pip install pandas openpyxl

COLONNES CORRECTES DU FICHIER SAGE :
  A (0)  : Commercial ou Code famille
  D (4)  : Intitulé famille
  M (12) : CA HT Net
  N (13) : % Remise
  Q (16) : Marge
  R (17) : % Marge sur CA
"""

import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ==============================
# CONFIG
# ==============================

FICHIER_SOURCE_DEFAUT = "STATCOLL-FAM.xlsx"

STYLES = {
    "title": {
        "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
        "fill": PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    },
    "header": {
        "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    },
    "kpi_header": {
        "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "fill": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    },
    "data_font": Font(name="Calibri", size=10),
    "kpi_font": Font(name="Calibri", size=10, bold=True),
    "border": Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    ),
}

COLONNES_LARGEURS = {
    "A": 22,  # Commercial
    "B": 16,  # Code Famille
    "C": 28,  # Intitulé
    "D": 16,  # CA HT Net
    "E": 14,  # % Remise
    "F": 16,  # Marge
    "G": 16,  # % Marge/CA
}

COLONNES_LARGEURS_KPI = {
    "A": 25,  # Commercial ou Famille
    "B": 16,  # Nombre de ventes
    "C": 18,  # CA
    "D": 18,  # Marge
    "E": 18,  # Profitabilité
}


# ==============================
# FONCTIONS
# ==============================

def afficher_titre():
    print("=" * 70)
    print("TRAITEMENT STATCOLL FAM - SAGE 100 (Avec KPI)")
    print("=" * 70)


def extraire_periode(wb):
    """
    Extrait les dates de période depuis la feuille Sage.
    Q2 = date début, Q3 = date fin.
    """
    ws = wb["Statistiques collaborateurs"]

    date_debut = ws["Q2"].value
    date_fin = ws["Q3"].value

    if isinstance(date_debut, datetime):
        date_debut_str = date_debut.strftime("%d/%m/%Y")
    else:
        date_debut_str = str(date_debut) if date_debut else "N/A"

    if isinstance(date_fin, datetime):
        date_fin_str = date_fin.strftime("%d/%m/%Y")
    else:
        date_fin_str = str(date_fin) if date_fin else "N/A"

    periode = f"{date_debut_str} au {date_fin_str}"
    return date_debut_str, date_fin_str, periode


def extraire_donnees(fichier_source):
    """
    Extrait les données en tenant compte de la structure réelle.
    """
    df = pd.read_excel(fichier_source, header=None)

    data_clean = []
    commercial_courant = None

    for _, row in df.iterrows():
        col_a = row[0]   # Commercial ou code famille
        col_d = row[4]   # Intitulé
        col_m = row[12]  # CA HT Net
        col_n = row[13]  # % Remise
        col_q = row[16]  # Marge
        col_r = row[17]  # % Marge sur CA

        # ligne vide -> on saute
        if all(pd.isna(v) for v in [col_a, col_d, col_m, col_n, col_q, col_r]):
            continue

        # Détection d'un commercial (nom seul en colonne A)
        if pd.notna(col_a) and isinstance(col_a, str):
            val_a = col_a.strip()

            # Lignes "Total ..." : on ignore pour le commercial
            if val_a.lower().startswith("total"):
                continue

            # Lignes de type "ABDESSAMED EL ORF", "COMPTOIR", etc.
            if not val_a.upper().startswith("F") and val_a.upper() not in ("ESCOMPTE",):
                commercial_courant = val_a
                continue

        # Détection des lignes de détail (codes famille)
        code_famille = None

        if pd.notna(col_a) and isinstance(col_a, str):
            val_a = col_a.strip()
            if val_a.upper().startswith("F"):
                code_famille = val_a
            elif val_a.upper() == "ESCOMPTE":
                code_famille = "ESCOMPTE"

        if code_famille is None:
            continue

        # Nettoyage / conversion numériques
        ca = float(col_m) if isinstance(col_m, (int, float)) else None
        remise = float(col_n) if isinstance(col_n, (int, float)) else None
        marge = float(col_q) if isinstance(col_q, (int, float)) else None
        pct_marge = float(col_r) if isinstance(col_r, (int, float)) else None

        intit = str(col_d).strip() if pd.notna(col_d) else ""

        data_clean.append(
            {
                "Commercial": commercial_courant,
                "Code Famille": code_famille,
                "Intitulé": intit,
                "CA HT Net": ca,
                "% Remise": remise,
                "Marge": marge,
                "% Marge/CA": pct_marge,
            }
        )

    return pd.DataFrame(data_clean)


def creer_feuille_statistiques(wb_new, df_clean, periode):
    """
    Crée la première feuille "Statistiques" avec le détail complet.
    """
    ws_new = wb_new.active
    ws_new.title = "Statistiques"

    # Titre
    ws_new.merge_cells("A1:G1")
    c_title = ws_new["A1"]
    c_title.value = f"Statistiques Collaborateurs/Famille - Période: {periode}"
    c_title.font = STYLES["title"]["font"]
    c_title.fill = STYLES["title"]["fill"]
    c_title.alignment = STYLES["title"]["alignment"]
    ws_new.row_dimensions[1].height = 32

    # Info
    ws_new.merge_cells("A2:G2")
    c_info = ws_new["A2"]
    c_info.value = "BBM AGRI | Tenue de compte en MAD"
    c_info.font = Font(name="Calibri", size=10, italic=True)
    c_info.alignment = Alignment(horizontal="left", vertical="center")
    ws_new.row_dimensions[2].height = 18

    ws_new.row_dimensions[3].height = 8  # ligne vide

    # En-têtes
    headers = [
        "Commercial",
        "Code Famille",
        "Intitulé",
        "CA HT Net",
        "% Remise",
        "Marge",
        "% Marge/CA",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws_new.cell(row=4, column=col)
        cell.value = header
        cell.font = STYLES["header"]["font"]
        cell.fill = STYLES["header"]["fill"]
        cell.alignment = STYLES["header"]["alignment"]
        cell.border = STYLES["border"]
    ws_new.row_dimensions[4].height = 24

    # Données
    for row_idx, (_, row) in enumerate(df_clean.iterrows(), start=5):
        values = row.tolist()
        for col_idx, value in enumerate(values, start=1):
            cell = ws_new.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = STYLES["data_font"]
            cell.border = STYLES["border"]

            if col_idx in (1, 2, 3):  # texte
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:  # numériques
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, (int, float)):
                    if col_idx in (5, 7):  # % Remise, % Marge/CA
                        cell.number_format = "0.00%"
                    else:
                        cell.number_format = "#,##0.00"

        ws_new.row_dimensions[row_idx].height = 18

    # Largeurs colonnes
    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws_new.column_dimensions[col_letter].width = COLONNES_LARGEURS[col_letter]

    # Pied de page
    last_row = len(df_clean) + 5
    ws_new.merge_cells(f"A{last_row + 2}:G{last_row + 2}")
    c_footer = ws_new[f"A{last_row + 2}"]
    c_footer.value = (
        f"Rapport généré le {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')} | "
        f"{len(df_clean)} enregistrement(s)"
    )
    c_footer.font = Font(name="Calibri", size=9, italic=True, color="666666")
    c_footer.alignment = Alignment(horizontal="right", vertical="center")


def _ecrire_tableau_double_entree(ws, start_row, title, df_pivot, format_type="nombre"):
    """
    Écrit un tableau à double entrée (Commerciaux × Familles) avec totaux.
    """
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=df_pivot.shape[1] + 2)
    c_title = ws.cell(row=start_row, column=1)
    c_title.value = title
    c_title.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c_title.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    c_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 24

    row = start_row + 1

    # En-têtes colonnes
    ws.cell(row=row, column=1, value="Commercial")
    ws.cell(row=row, column=1).font = STYLES["kpi_header"]["font"]
    ws.cell(row=row, column=1).fill = STYLES["kpi_header"]["fill"]
    ws.cell(row=row, column=1).alignment = STYLES["kpi_header"]["alignment"]
    ws.cell(row=row, column=1).border = STYLES["border"]

    for col_idx, fam in enumerate(df_pivot.columns, start=2):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = fam
        cell.font = STYLES["kpi_header"]["font"]
        cell.fill = STYLES["kpi_header"]["fill"]
        cell.alignment = STYLES["kpi_header"]["alignment"]
        cell.border = STYLES["border"]

    cell_tot = ws.cell(row=row, column=df_pivot.shape[1] + 2)
    cell_tot.value = "Total"
    cell_tot.font = STYLES["kpi_header"]["font"]
    cell_tot.fill = STYLES["kpi_header"]["fill"]
    cell_tot.alignment = STYLES["kpi_header"]["alignment"]
    cell_tot.border = STYLES["border"]

    ws.row_dimensions[row].height = 20
    row += 1

    # Lignes de données
    for _, (commercial, row_values) in enumerate(df_pivot.iterrows()):
        c = ws.cell(row=row, column=1)
        c.value = commercial
        c.font = STYLES["data_font"]
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = STYLES["border"]

        row_total = 0
        for col_idx, fam in enumerate(df_pivot.columns, start=2):
            val = row_values[fam]
            cell = ws.cell(row=row, column=col_idx)
            cell.value = float(val) if pd.notna(val) else 0
            cell.font = STYLES["data_font"]
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = STYLES["border"]

            if format_type == "montant":
                cell.number_format = "#,##0.00"
            elif format_type == "pourcentage":
                cell.number_format = "0.00%"
            else:
                cell.number_format = "0"

            row_total += cell.value

        c_tot = ws.cell(row=row, column=df_pivot.shape[1] + 2)
        c_tot.value = row_total
        c_tot.font = STYLES["kpi_font"]
        c_tot.alignment = Alignment(horizontal="right", vertical="center")
        c_tot.border = STYLES["border"]
        if format_type == "montant":
            c_tot.number_format = "#,##0.00"
        elif format_type == "pourcentage":
            c_tot.number_format = "0.00%"
        else:
            c_tot.number_format = "0"

        ws.row_dimensions[row].height = 18
        row += 1

    # Totaux colonnes
    c_label = ws.cell(row=row, column=1)
    c_label.value = "Total"
    c_label.font = STYLES["kpi_font"]
    c_label.alignment = Alignment(horizontal="left", vertical="center")
    c_label.border = STYLES["border"]

    for col_idx, fam in enumerate(df_pivot.columns, start=2):
        col_sum = df_pivot[fam].sum()
        c = ws.cell(row=row, column=col_idx)
        c.value = float(col_sum)
        c.font = STYLES["kpi_font"]
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = STYLES["border"]
        if format_type == "montant":
            c.number_format = "#,##0.00"
        elif format_type == "pourcentage":
            c.number_format = "0.00%"
        else:
            c.number_format = "0"

    grand_total = df_pivot.to_numpy().sum()
    c_gt = ws.cell(row=row, column=df_pivot.shape[1] + 2)
    c_gt.value = float(grand_total)
    c_gt.font = STYLES["kpi_font"]
    c_gt.alignment = Alignment(horizontal="right", vertical="center")
    c_gt.border = STYLES["border"]
    if format_type == "montant":
        c_gt.number_format = "#,##0.00"
    elif format_type == "pourcentage":
        c_gt.number_format = "0.00%"
    else:
        c_gt.number_format = "0"

    ws.row_dimensions[row].height = 20

    return row + 2


def creer_feuille_kpi(wb_new, df_clean, periode):
    """
    Crée la feuille KPI avec résumés + 4 tableaux Commerciaux × Familles.
    """
    ws_kpi = wb_new.create_sheet("KPI")

    # Titre
    ws_kpi.merge_cells("A1:E1")
    c_title = ws_kpi["A1"]
    c_title.value = f"KPI & Tableau de Bord - Période: {periode}"
    c_title.font = STYLES["title"]["font"]
    c_title.fill = STYLES["title"]["fill"]
    c_title.alignment = STYLES["title"]["alignment"]
    ws_kpi.row_dimensions[1].height = 32

    # ========== RÉSUMÉ PAR COMMERCIAL ==========
    ligne_actuelle = 3

    ws_kpi.merge_cells(f"A{ligne_actuelle}:E{ligne_actuelle}")
    c_section = ws_kpi[f"A{ligne_actuelle}"]
    c_section.value = "RÉSUMÉ PAR COMMERCIAL"
    c_section.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c_section.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    c_section.alignment = Alignment(horizontal="center", vertical="center")
    ws_kpi.row_dimensions[ligne_actuelle].height = 24

    ligne_actuelle += 1

    headers_kpi = ["Commercial", "Nbre Ventes", "CA Total", "Marge Total", "Profitabilité"]
    for col, header in enumerate(headers_kpi, start=1):
        cell = ws_kpi.cell(row=ligne_actuelle, column=col)
        cell.value = header
        cell.font = STYLES["kpi_header"]["font"]
        cell.fill = STYLES["kpi_header"]["fill"]
        cell.alignment = STYLES["kpi_header"]["alignment"]
        cell.border = STYLES["border"]
    ws_kpi.row_dimensions[ligne_actuelle].height = 22

    ligne_actuelle += 1

    summary_commercial = df_clean.groupby("Commercial").agg({
        "Code Famille": "count",
        "CA HT Net": "sum",
        "Marge": "sum"
    }).reset_index()
    summary_commercial.columns = ["Commercial", "Nbre Ventes", "CA Total", "Marge Total"]
    summary_commercial["Profitabilité"] = (
        summary_commercial["Marge Total"] / summary_commercial["CA Total"]
    ).fillna(0)

    for _, row in summary_commercial.iterrows():
        for col_idx, col_name in enumerate(["Commercial", "Nbre Ventes", "CA Total", "Marge Total", "Profitabilité"], start=1):
            cell = ws_kpi.cell(row=ligne_actuelle, column=col_idx)
            value = row[col_name]
            cell.value = value
            cell.font = STYLES["data_font"]
            cell.border = STYLES["border"]

            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col_idx == 2:
                    cell.number_format = "0"
                elif col_idx in (3, 4):
                    cell.number_format = "#,##0.00"
                elif col_idx == 5:
                    cell.number_format = "0.00%"

        ws_kpi.row_dimensions[ligne_actuelle].height = 18
        ligne_actuelle += 1

    ligne_actuelle += 1
    cell_total = ws_kpi[f"A{ligne_actuelle}"]
    cell_total.value = "TOTAL GÉNÉRAL"
    cell_total.font = STYLES["kpi_font"]
    cell_total.alignment = Alignment(horizontal="left", vertical="center")
    cell_total.border = STYLES["border"]

    total_ventes = len(df_clean)
    total_ca = df_clean["CA HT Net"].sum()
    total_marge = df_clean["Marge"].sum()
    total_profitabilite = (total_marge / total_ca) if total_ca else 0

    for col_idx, value in enumerate([total_ventes, total_ca, total_marge, total_profitabilite], start=2):
        cell = ws_kpi.cell(row=ligne_actuelle, column=col_idx)
        cell.value = value
        cell.font = STYLES["kpi_font"]
        cell.border = STYLES["border"]
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if col_idx == 2:
            cell.number_format = "0"
        elif col_idx in (3, 4):
            cell.number_format = "#,##0.00"
        elif col_idx == 5:
            cell.number_format = "0.00%"

    ws_kpi.row_dimensions[ligne_actuelle].height = 20

    # ========== RÉSUMÉ PAR FAMILLE ==========
    ligne_actuelle += 3

    ws_kpi.merge_cells(f"A{ligne_actuelle}:E{ligne_actuelle}")
    c_section = ws_kpi[f"A{ligne_actuelle}"]
    c_section.value = "RÉSUMÉ PAR FAMILLE"
    c_section.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c_section.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    c_section.alignment = Alignment(horizontal="center", vertical="center")
    ws_kpi.row_dimensions[ligne_actuelle].height = 24

    ligne_actuelle += 1

    for col, header in enumerate(headers_kpi, start=1):
        header_famille = "Famille" if col == 1 else header
        cell = ws_kpi.cell(row=ligne_actuelle, column=col)
        cell.value = header_famille
        cell.font = STYLES["kpi_header"]["font"]
        cell.fill = STYLES["kpi_header"]["fill"]
        cell.alignment = STYLES["kpi_header"]["alignment"]
        cell.border = STYLES["border"]
    ws_kpi.row_dimensions[ligne_actuelle].height = 22

    ligne_actuelle += 1

    # Résumé par famille (corrigé pour éviter les colonnes multi-niveaux)
    group = df_clean.groupby("Code Famille", as_index=False)
    summary_famille = group.agg({
        "Intitulé": "first",
        "CA HT Net": "sum",
        "Marge": "sum"
    })
    summary_famille["Nbre Ventes"] = group["CA HT Net"].size().values
    summary_famille.rename(columns={
        "CA HT Net": "CA Total",
        "Marge": "Marge Total"
    }, inplace=True)
    summary_famille["Profitabilité"] = (
        summary_famille["Marge Total"] / summary_famille["CA Total"]
    ).fillna(0)
    summary_famille["Famille"] = summary_famille["Code Famille"] + " - " + summary_famille["Intitulé"]

    for _, row in summary_famille.iterrows():
        cell_col1 = ws_kpi.cell(row=ligne_actuelle, column=1)
        cell_col1.value = row["Famille"]
        cell_col1.font = STYLES["data_font"]
        cell_col1.border = STYLES["border"]
        cell_col1.alignment = Alignment(horizontal="left", vertical="center")

        for col_idx, col_name in enumerate(["Nbre Ventes", "CA Total", "Marge Total", "Profitabilité"], start=2):
            cell = ws_kpi.cell(row=ligne_actuelle, column=col_idx)
            value = row[col_name]
            cell.value = value
            cell.font = STYLES["data_font"]
            cell.border = STYLES["border"]
            cell.alignment = Alignment(horizontal="right", vertical="center")

            if col_idx == 2:
                cell.number_format = "0"
            elif col_idx in (3, 4):
                cell.number_format = "#,##0.00"
            elif col_idx == 5:
                cell.number_format = "0.00%"

        ws_kpi.row_dimensions[ligne_actuelle].height = 18
        ligne_actuelle += 1

    for col_letter in ["A", "B", "C", "D", "E"]:
        ws_kpi.column_dimensions[col_letter].width = COLONNES_LARGEURS_KPI[col_letter]

    # ========== TABLEAUX DOUBLE ENTRÉE COMMERCIAUX × FAMILLES ==========
    ligne_actuelle += 3

    pivot_nbre = pd.pivot_table(
        df_clean,
        index="Commercial",
        columns="Code Famille",
        values="Code Famille",
        aggfunc="count",
        fill_value=0
    )
    ligne_actuelle = _ecrire_tableau_double_entree(
        ws_kpi, ligne_actuelle,
        "TABLEAU NOMBRE DE VENTES (Commerciaux × Familles)",
        pivot_nbre,
        format_type="nombre"
    )

    pivot_ca = pd.pivot_table(
        df_clean,
        index="Commercial",
        columns="Code Famille",
        values="CA HT Net",
        aggfunc="sum",
        fill_value=0
    )
    ligne_actuelle = _ecrire_tableau_double_entree(
        ws_kpi, ligne_actuelle,
        "TABLEAU CA HT NET (Commerciaux × Familles)",
        pivot_ca,
        format_type="montant"
    )

    pivot_marge = pd.pivot_table(
        df_clean,
        index="Commercial",
        columns="Code Famille",
        values="Marge",
        aggfunc="sum",
        fill_value=0
    )
    ligne_actuelle = _ecrire_tableau_double_entree(
        ws_kpi, ligne_actuelle,
        "TABLEAU MARGE TOTALE (Commerciaux × Familles)",
        pivot_marge,
        format_type="montant"
    )

    profit_matrix = pivot_marge.copy()
    profit_matrix = profit_matrix / pivot_ca.replace(0, float("nan"))
    profit_matrix = profit_matrix.fillna(0)

    _ = _ecrire_tableau_double_entree(
        ws_kpi, ligne_actuelle,
        "TABLEAU PROFITABILITÉ (Marge / CA) (Commerciaux × Familles)",
        profit_matrix,
        format_type="pourcentage"
    )


def creer_excel_structure(df_clean, periode, date_debut_str, date_fin_str):
    output_name = (
        f"STATCOLL-FAM_PERIODE_{date_debut_str.replace('/', '-')}"
        f"_au_{date_fin_str.replace('/', '-')}.xlsx"
    )

    wb_new = openpyxl.Workbook()
    creer_feuille_statistiques(wb_new, df_clean, periode)
    creer_feuille_kpi(wb_new, df_clean, periode)

    wb_new.save(output_name)
    return output_name


# ==============================
# MAIN
# ==============================

def main():
    afficher_titre()

    if len(sys.argv) > 1:
        fichier_source = sys.argv[1]
    else:
        fichier_source = FICHIER_SOURCE_DEFAUT

    if not os.path.exists(fichier_source):
        print(f"❌ Fichier introuvable : {fichier_source}")
        print("   Place le fichier dans le même dossier ou passe le chemin en argument.")
        sys.exit(1)

    print(f"📂 Fichier source : {fichier_source}")

    try:
        print("⏳ Lecture du classeur...")
        wb = openpyxl.load_workbook(fichier_source)

        print("📅 Extraction de la période...")
        date_debut_str, date_fin_str, periode = extraire_periode(wb)
        print(f"   Période : {periode}")

        print("📊 Extraction des données...")
        df_clean = extraire_donnees(fichier_source)
        print(f"   Lignes extraites : {len(df_clean)}")

        print("🔧 Génération du fichier structuré avec KPI...")
        output_file = creer_excel_structure(df_clean, periode, date_debut_str, date_fin_str)

        print("\n✅ Terminé.")
        print(f"   Fichier généré : {output_file}")
        print(f"   Période : {periode}")
        print(f"   Lignes de données : {len(df_clean)}")
        print(f"   Feuilles : Statistiques + KPI (avec 4 tableaux croisés)")

    except Exception as e:
        print("\n❌ ERREUR pendant le traitement :")
        print(e)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
