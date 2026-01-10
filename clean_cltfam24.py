#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script pour nettoyer les fichiers CLTFAM*.xlsx et générer un Excel propre
Usage: python clean_cltfam24.py [fichier_source.xlsx]
Si aucun fichier n'est spécifié, traite CLTFAM24.xlsx par défaut
"""

import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Configuration par défaut
FICHIER_SOURCE_DEFAUT = "CLTFAM24.xlsx"

STYLES = {
    "title": {
        "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
        "fill": PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    },
    "header": {
        "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    },
    "data_font": Font(name="Calibri", size=10),
    "border": Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    ),
}

COLONNES_LARGEURS = {
    "A": 22,  # Commercial
    "B": 16,  # Code Famille
    "C": 28,  # Intitulé
    "D": 16,  # CA HT Net
    "E": 14,  # % Remise
    "F": 16,  # Marge
    "G": 16,  # % Marge/CA
}


def extraire_donnees(fichier_source):
    """
    Extrait et nettoie les données du fichier Excel CLTFAM*.
    Structure réelle détectée :
    - Colonne A (0) : Code commercial (C00001, C00003...) ou Code famille (F002, F003...)
    - Colonne D (3) : Nom commercial ou Intitulé famille
    - Colonne L (11) : CA HT Net
    - Colonne N (13) : % Remise (décimal, ex: 0.0035 = 0.35%)
    - Colonne P (15) : Marge
    - Colonne R (17) : % Marge sur CA
    """
    df = pd.read_excel(fichier_source, header=None)
    
    data_clean = []
    commercial_courant = None
    nom_commercial_courant = None
    
    for _, row in df.iterrows():
        col_a = row[0]   # Code commercial ou code famille
        col_d = row[3]   # Nom commercial ou intitulé famille
        col_l = row[11] if len(row) > 11 else None  # CA HT Net
        col_n = row[13] if len(row) > 13 else None  # % Remise
        col_p = row[15] if len(row) > 15 else None  # Marge
        col_r = row[17] if len(row) > 17 else None  # % Marge sur CA
        
        # Ligne vide -> on saute
        if pd.isna(col_a):
            continue
        
        if not isinstance(col_a, str):
            continue
        
        val_a = str(col_a).strip()
        
        # Ignorer les lignes d'en-tête
        if val_a.lower().startswith("total") or val_a.lower().startswith("sage") or val_a == "BBM AGRI":
            continue
        
        # Détection d'un code commercial (C00001, C00003, etc.)
        if val_a.upper().startswith("C") and len(val_a) >= 5 and val_a[1:].isdigit():
            commercial_courant = val_a
            # Le nom du commercial est dans la colonne D (3)
            if pd.notna(col_d) and isinstance(col_d, str):
                nom_commercial_courant = str(col_d).strip()
            else:
                nom_commercial_courant = val_a
            continue
        
        # Détection des lignes de détail (codes famille F002, F003, etc.)
        code_famille = None
        
        if val_a.upper().startswith("F") and len(val_a) >= 2:
            code_famille = val_a
        elif val_a.upper() == "ESCOMPTE":
            code_famille = "ESCOMPTE"
        
        if code_famille is None:
            continue
        
        # Nettoyage / conversion numériques
        ca = None
        remise = None
        marge = None
        pct_marge = None
        
        if pd.notna(col_l):
            try:
                ca = float(col_l) if isinstance(col_l, (int, float)) else float(str(col_l).replace(",", "."))
            except (ValueError, TypeError):
                ca = None
        
        if pd.notna(col_n):
            try:
                # Le % remise est en décimal (0.0035 = 0.35%)
                remise = float(col_n) if isinstance(col_n, (int, float)) else float(str(col_n).replace(",", "."))
            except (ValueError, TypeError):
                remise = None
        
        if pd.notna(col_p):
            try:
                marge = float(col_p) if isinstance(col_p, (int, float)) else float(str(col_p).replace(",", "."))
            except (ValueError, TypeError):
                marge = None
        
        if pd.notna(col_r):
            try:
                pct_marge = float(col_r) if isinstance(col_r, (int, float)) else float(str(col_r).replace(",", "."))
            except (ValueError, TypeError):
                pct_marge = None
        
        # Ignorer les lignes sans CA ni Marge
        if ca is None and marge is None:
            continue
        
        # L'intitulé de la famille est dans la colonne D (3)
        intit = str(col_d).strip() if pd.notna(col_d) and isinstance(col_d, str) else ""
        
        data_clean.append({
            "Commercial": nom_commercial_courant or commercial_courant,
            "Code Famille": code_famille,
            "Intitulé": intit,
            "CA HT Net": ca,
            "% Remise": remise,
            "Marge": marge,
            "% Marge/CA": pct_marge,
        })
    
    return pd.DataFrame(data_clean)


def creer_excel_clean(df_clean, fichier_sortie, nom_source):
    """
    Crée un fichier Excel propre avec les données nettoyées.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Données nettoyées"
    
    # Titre
    ws.merge_cells("A1:G1")
    c_title = ws["A1"]
    # Extraire le nom du fichier sans extension pour le titre
    nom_base = os.path.splitext(os.path.basename(nom_source))[0]
    c_title.value = f"{nom_base} - Données nettoyées"
    c_title.font = STYLES["title"]["font"]
    c_title.fill = STYLES["title"]["fill"]
    c_title.alignment = STYLES["title"]["alignment"]
    ws.row_dimensions[1].height = 32
    
    # Info
    ws.merge_cells("A2:G2")
    c_info = ws["A2"]
    c_info.value = f"BBM AGRI | Fichier généré le {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}"
    c_info.font = Font(name="Calibri", size=10, italic=True)
    c_info.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18
    
    ws.row_dimensions[3].height = 8  # ligne vide
    
    # En-têtes
    headers = [
        "Commercial",
        "Code Famille",
        "Intitulé",
        "CA HT Net",
        "% Remise",
        "Marge",
        "% Marge/CA",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = STYLES["header"]["font"]
        cell.fill = STYLES["header"]["fill"]
        cell.alignment = STYLES["header"]["alignment"]
        cell.border = STYLES["border"]
    ws.row_dimensions[4].height = 24
    
    # Données
    for row_idx, (_, row) in enumerate(df_clean.iterrows(), start=5):
        values = row.tolist()
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = STYLES["data_font"]
            cell.border = STYLES["border"]
            
            if col_idx in (1, 2, 3):  # texte
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:  # numériques
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, (int, float)) and pd.notna(value):
                    if col_idx in (5, 7):  # % Remise, % Marge/CA
                        cell.number_format = "0.00%"
                    else:
                        cell.number_format = "#,##0.00"
        
        ws.row_dimensions[row_idx].height = 18
    
    # Largeurs colonnes
    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = COLONNES_LARGEURS[col_letter]
    
    # Pied de page
    last_row = len(df_clean) + 5
    ws.merge_cells(f"A{last_row + 2}:G{last_row + 2}")
    c_footer = ws[f"A{last_row + 2}"]
    c_footer.value = f"{len(df_clean)} enregistrement(s)"
    c_footer.font = Font(name="Calibri", size=9, italic=True, color="666666")
    c_footer.alignment = Alignment(horizontal="right", vertical="center")
    
    wb.save(fichier_sortie)
    return fichier_sortie


def main():
    print("=" * 70)
    print("NETTOYAGE CLTFAM*.xlsx")
    print("=" * 70)
    
    # Récupérer le fichier source depuis les arguments ou utiliser le défaut
    if len(sys.argv) > 1:
        fichier_source = sys.argv[1]
    else:
        fichier_source = FICHIER_SOURCE_DEFAUT
    
    if not os.path.exists(fichier_source):
        print(f"ERREUR: Fichier introuvable : {fichier_source}")
        print(f"   Usage: python clean_cltfam24.py [fichier_source.xlsx]")
        sys.exit(1)
    
    # Générer le nom du fichier de sortie
    nom_base = os.path.splitext(os.path.basename(fichier_source))[0]
    fichier_sortie = f"{nom_base}_clean.xlsx"
    
    print(f"Fichier source : {fichier_source}")
    print(f"Fichier sortie : {fichier_sortie}")
    
    try:
        print("Lecture du fichier...")
        df_clean = extraire_donnees(fichier_source)
        print(f"  Lignes extraites : {len(df_clean)}")
        
        if len(df_clean) == 0:
            print("ERREUR: Aucune donnee extraite. Verifiez le format du fichier.")
            sys.exit(1)
        
        print(f"Generation du fichier Excel propre...")
        fichier_sortie = creer_excel_clean(df_clean, fichier_sortie, fichier_source)
        
        print("\nTERMINE")
        print(f"  Fichier genere : {fichier_sortie}")
        print(f"  Lignes de donnees : {len(df_clean)}")
        
    except Exception as e:
        print(f"\nERREUR pendant le traitement :")
        print(e)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
