#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
STATCOLL ART – Traitement Sage 100 (2 feuilles Excel)

- Lecture d’un fichier Sage 100 : Statistiques Collaborateurs / Article
- Génération d’un fichier Excel structuré :
    1. Statistiques détaillées par article
    2. KPI par commercial
"""

# ============================================================
# IMPORTS
# ============================================================

import sys
import os
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================================
# CONFIGURATION GLOBALE
# ============================================================

FICHIER_SOURCE_DEFAUT = "STATCOLL-ART.xlsx"

STYLES = {
    "title": {
        "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
        "fill": PatternFill("solid", fgColor="1F4E78"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    },
    "header": {
        "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "fill": PatternFill("solid", fgColor="4472C4"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    },
    "data_font": Font(name="Calibri", size=10),
    "border": Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    ),
}

COLONNES_LARGEURS = {
    "A": 22,
    "B": 16,
    "C": 40,
    "D": 16,
    "E": 14,
    "F": 16,
    "G": 16,
}


# ============================================================
# AFFICHAGE & UTILITAIRES
# ============================================================

def afficher_titre():
    print("=" * 70)
    print("TRAITEMENT STATCOLL ART – SAGE 100")
    print("=" * 70)


def extraire_periode(wb):
    """Extraction des dates de période depuis la feuille Sage."""
    ws = wb["Statistiques collaborateurs"]

    def _fmt(date):
        return date.strftime("%d/%m/%Y") if isinstance(date, datetime) else str(date)

    debut = ws["Q2"].value
    fin = ws["Q3"].value

    debut_str = _fmt(debut) if debut else "N/A"
    fin_str = _fmt(fin) if fin else "N/A"

    return debut_str, fin_str, f"{debut_str} au {fin_str}"


# ============================================================
# EXTRACTION DES DONNÉES
# ============================================================

def extraire_donnees(fichier_source):
    """
    Lecture de la feuille Sage et normalisation des données.
    """
    df = pd.read_excel(fichier_source, sheet_name=0, header=None)

    data = []
    commercial = None

    for idx, row in df.iterrows():

        if idx < 13:
            continue

        col_a, col_e = row[0], row[4]

        if pd.notna(col_a) and pd.isna(col_e):
            val = str(col_a).strip()
            if not val.lower().startswith("total"):
                commercial = val
            continue

        if pd.notna(col_a) and pd.notna(col_e):
            if pd.isna(row[12]) and pd.isna(row[16]):
                continue

            data.append({
                "Commercial": commercial,
                "Code Article": str(col_a).strip(),
                "Désignation": str(col_e).strip(),
                "CA HT Net": float(row[12]) if pd.notna(row[12]) else 0,
                "% Remise": float(row[13]) if pd.notna(row[13]) else 0,
                "Marge": float(row[16]) if pd.notna(row[16]) else 0,
                "% Marge/CA": float(row[17]) if pd.notna(row[17]) else 0,
            })

    return pd.DataFrame(data)


# ============================================================
# CALCULS KPI
# ============================================================

def calculer_kpi(df):
    """Calcul des indicateurs par commercial."""
    kpi = (
        df.groupby("Commercial", as_index=False)
          .agg({
              "CA HT Net": "sum",
              "Marge": "sum",
              "Code Article": "count"
          })
          .rename(columns={"Code Article": "Nb Articles"})
    )

    kpi["% Marge/CA Moyen"] = (
        kpi["Marge"] / kpi["CA HT Net"].replace(0, pd.NA) * 100
    ).fillna(0)

    return kpi.sort_values("CA HT Net", ascending=False).reset_index(drop=True)


# ============================================================
# GÉNÉRATION EXCEL
# ============================================================

def creer_feuille_statistiques(wb, df, periode):
    ws = wb.active
    ws.title = "Statistiques"

    ws.merge_cells("A1:G1")
    ws["A1"].value = f"Statistiques Collaborateurs / Article – {periode}"
    ws["A1"].font = STYLES["title"]["font"]
    ws["A1"].fill = STYLES["title"]["fill"]
    ws["A1"].alignment = STYLES["title"]["alignment"]
    ws.row_dimensions[1].height = 32

    headers = list(df.columns)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = STYLES["header"]["font"]
        cell.fill = STYLES["header"]["fill"]
        cell.alignment = STYLES["header"]["alignment"]
        cell.border = STYLES["border"]

    for i, (_, row) in enumerate(df.iterrows(), start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = STYLES["data_font"]
            cell.border = STYLES["border"]

    for col, width in COLONNES_LARGEURS.items():
        ws.column_dimensions[col].width = width


def creer_feuille_kpi(wb, kpi):
    ws = wb.create_sheet("KPI")

    headers = kpi.columns.tolist()

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = STYLES["header"]["font"]
        cell.fill = STYLES["header"]["fill"]
        cell.alignment = STYLES["header"]["alignment"]
        cell.border = STYLES["border"]

    for i, (_, row) in enumerate(kpi.iterrows(), start=5):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)


def generer_excel(df, periode, d1, d2):
    output = f"STATCOLL-ART_PERIODE_{d1.replace('/', '-')}_au_{d2.replace('/', '-')}.xlsx"
    wb = openpyxl.Workbook()

    creer_feuille_statistiques(wb, df, periode)
    creer_feuille_kpi(wb, calculer_kpi(df))

    wb.save(output)
    return output


# ============================================================
# MAIN
# ============================================================

def main():
    afficher_titre()

    fichier = sys.argv[1] if len(sys.argv) > 1 else FICHIER_SOURCE_DEFAUT

    if not os.path.exists(fichier):
        print(f"❌ Fichier introuvable : {fichier}")
        sys.exit(1)

    wb = openpyxl.load_workbook(fichier)
    d1, d2, periode = extraire_periode(wb)

    df = extraire_donnees(fichier)
    output = generer_excel(df, periode, d1, d2)

    print(f"✅ Fichier généré : {output}")


if __name__ == "__main__":
    main()
